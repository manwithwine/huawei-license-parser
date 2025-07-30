import os
import re
from openpyxl import Workbook, load_workbook


def find_esn_in_files(folder_path):
    esn_data = {}
    esn_pattern = re.compile(r'<ESN>([A-Za-z0-9]+)</ESN>')

    for filename in os.listdir(folder_path):
        if filename.endswith('.xml'):
            file_path = os.path.join(folder_path, filename)
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    content = file.read()
                    match = esn_pattern.search(content)
                    if match:
                        esn = match.group(1)
                        esn_data[esn] = filename
            except UnicodeDecodeError:
                try:
                    # Try another encoding if utf-8 doesn't work
                    with open(file_path, 'r', encoding='windows-1251') as file:
                        content = file.read()
                        match = esn_pattern.search(content)
                        if match:
                            esn = match.group(1)
                            esn_data[esn] = filename
                except Exception as e:
                    print(f"Error {filename}: {str(e)}")
            except Exception as e:
                print(f"Error {filename}: {str(e)}")
    return esn_data


def process_table(input_xlsx, esn_data, output_xlsx):
    # Uploading XLSX file
    wb = load_workbook(input_xlsx)
    ws = wb.active

    # Add a header for the third column if it doesn't exist
    if ws.max_column < 3:
        ws.cell(row=1, column=3, value="Source File")
    else:
        if not ws.cell(row=1, column=3).value:
            ws.cell(row=1, column=3, value="Source File")

    # Fill in the third column
    for row in range(2, ws.max_row + 1):  # Skip the title
        serial = ws.cell(row=row, column=2).value  # S/N in the second column
        if serial and str(serial) in esn_data:
            ws.cell(row=row, column=3, value=esn_data[str(serial)])

    # Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_xlsx)


if __name__ == "__main__":

    xml_folder = "license"  # Folder with xml files
    input_table = "table.xlsx"  # Input table (XLSX)
    output_table = "result.xlsx"  # Output table (XLSX)

    # Step 1: Find all ESN in xml files
    esn_dict = find_esn_in_files(xml_folder)
    print(f"Found {len(esn_dict)} ESN in xml files")

    # Step 2: Process the table
    process_table(input_table, esn_dict, output_table)
    print(f"Result saved in {output_table}")